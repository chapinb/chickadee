"""
XLSX File Parser
==================

Parse IP addresses from XLSX files. This will extract IP addresses stored as
values (not formulas) across all tabs within a spreadsheet.

"""

from openpyxl import load_workbook

from libchickadee.parsers import ParserBase, run_parser_from_cli

__author__ = 'Chapin Bryce'
__date__ = 20200107
__license__ = 'MIT Copyright 2020 Chapin Bryce'
__desc__ = '''Yet another GeoIP resolution tool.'''


class XLSXParser(ParserBase):
    """Class to extract IP addresses from XLSX workbooks."""
    def __init__(self, ignore_bogon=True):
        super().__init__(ignore_bogon)
        self.ips = {}

    def parse_file(self, file_entry, is_stream=False):
        """Parse xlsx contents. Must be a path to an existing XLSX workbook.
        Cannot parse from STDIN.

        Args:
            file_entry (str): Path to workbook to load.
            is_stream (bool): Unused argument, required for implementation.
                Does not change functionality.
        """
        if is_stream:
            raise NotImplementedError("Providing XLSX files as an input stream of data is not yet supported.")

        wb = load_workbook(file_entry)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (str, bytes)):
                        self.check_ips(cell.value)


if __name__ == '__main__':  # pragma: no cover
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('path', help="File or folder to parse")
    args = parser.parse_args()

    xl_parser = XLSXParser()
    run_parser_from_cli(args=args, parser_obj=xl_parser)
